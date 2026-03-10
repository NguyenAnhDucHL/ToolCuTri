import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os
import re
import io
import datetime
import tempfile
import random
import unicodedata
import subprocess

# Absolute path to the picker helper script (same directory as this file)
_PICKER_SCRIPT = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'picker.py')

# ─────────────────────────────────────────────────────────────────────────────
# Utility helpers
# ─────────────────────────────────────────────────────────────────────────────

ELECTION_DATE = datetime.date(2026, 3, 15)

# Dates for first-time 18-year-old voters (turned 18 between election dates)
DOB_18_START = datetime.date(2007, 3, 16)
DOB_18_END   = datetime.date(2008, 3, 15)

# Born before this date → older than 80 as of election day
DOB_ELDERLY  = datetime.date(1946, 3, 15)


def normalize_text(text: str) -> str:
    """Lowercase + strip + remove diacritics for fuzzy matching."""
    if not isinstance(text, str):
        text = str(text)
    text = unicodedata.normalize("NFD", text)
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")
    return text.lower().strip()


TOTAL_KEYWORDS = re.compile(
    r"t[oô]ng\s*(s[oô]|h[ooa]p|c[oô]ng|c[uư] tri)?",
    re.IGNORECASE | re.UNICODE,
)

def is_total_row_label(value) -> bool:
    """Return True if the cell value looks like a 'Tổng' label."""
    if not value:
        return False
    norm = normalize_text(str(value))
    return bool(TOTAL_KEYWORDS.search(norm))


def is_positive_int(value) -> bool:
    try:
        return float(value) > 0 and float(value) == int(float(value))
    except (TypeError, ValueError):
        return False


def extract_numeric(value):
    """Return int if value is a positive integer, else None."""
    try:
        v = float(value)
        if v > 0:
            return int(v)
    except (TypeError, ValueError):
        pass
    return None


# ─────────────────────────────────────────────────────────────────────────────
# Google Drive helpers
# ─────────────────────────────────────────────────────────────────────────────

def parse_gdrive_folder_id(url: str):
    """Extract folder ID from a Google Drive folder URL."""
    match = re.search(r"/folders/([a-zA-Z0-9_-]+)", url)
    if match:
        return match.group(1)
    return None


def parse_gdrive_file_id(url: str):
    """Extract file ID from a Google Drive file URL."""
    patterns = [
        r"/file/d/([a-zA-Z0-9_-]+)",
        r"id=([a-zA-Z0-9_-]+)",
        r"/d/([a-zA-Z0-9_-]+)",
    ]
    for p in patterns:
        m = re.search(p, url)
        if m:
            return m.group(1)
    return None


def download_gdrive_folder(folder_id: str, dest_dir: str) -> list:
    """Download all .xlsx files from a public Google Drive folder."""
    try:
        import gdown
        url = f"https://drive.google.com/drive/folders/{folder_id}"
        gdown.download_folder(url, output=dest_dir, quiet=True, use_cookies=False)
        files = []
        for root, _, filenames in os.walk(dest_dir):
            for fn in filenames:
                if fn.lower().endswith(".xlsx"):
                    files.append(os.path.join(root, fn))
        return files
    except Exception as e:
        raise RuntimeError(f"Không thể tải folder Google Drive: {e}")


def download_gdrive_file(file_id: str, dest_path: str):
    """Download a single file from Google Drive by ID."""
    try:
        import gdown
        url = f"https://drive.google.com/uc?id={file_id}"
        gdown.download(url, dest_path, quiet=True)
    except Exception as e:
        raise RuntimeError(f"Không thể tải file Google Drive: {e}")


# ─────────────────────────────────────────────────────────────────────────────
# Source file parsing
# ─────────────────────────────────────────────────────────────────────────────

# Regex to extract Tổng/Nam/Nữ from free-text sentences:
_TEXT_TOTAL_RE = re.compile(
    r'là[:\s]*(\d+)\s*người[^\n]*?[:\s]+(\d+)\s*[Nn]am[;,\s]+(\d+)\s*[Nn]ữ',
    re.IGNORECASE | re.UNICODE,
)
_TEXT_ALT_RE = re.compile(
    r'(\d+)\s*người[^\n]*?(\d+)\s*[Nn]am[;,\s]+(\d+)\s*[Nn]ữ',
    re.IGNORECASE | re.UNICODE
)
# Alt pattern for "Nam: X; Nữ: Y" format
_TEXT_NAM_NU_RE = re.compile(
    r'(\d+)\s*người(?:[,;\s]|trong đó)+[Nn]am[:\s]*(\d+)[;,\s]+[Nn]ữ[:\s]*(\d+)',
    re.IGNORECASE | re.UNICODE
)
# Regex to find cell references inside a formula like =...&B584&...&C584&...&D584&...
# Captures up to 4 cell refs that are likely to hold the numbers
_FORMULA_REF_RE = re.compile(r'&([A-Z]+\d+)&', re.IGNORECASE)


# ── Sheet priority groups (normalized names) ─────────────────────────────────
# Priority 1: Explicit 'Tổng hợp' summary sheets
_SUMMARY_SHEET_NAMES = {
    "tong hop cu tri", "tong hop", "tonghop", "tonghopcu tri"
}
# Priority 2: Total/aggregate sheets (may contain T/N/N in table or text)
_TOTAL_SHEET_NAMES = {
    "tong",          # Tổng
    "sheet tong",    # SHEET TỔNG (Khu phố Minh Khai)
}


def _norm(name: str) -> str:
    """Normalize a sheet name for comparison."""
    return normalize_text(name)


def get_candidate_sheets(wb: openpyxl.Workbook) -> list:
    """
    Return ALL worksheets in priority order to search for Tổng/Nam/Nữ data:
    1. Explicit summary sheets ('Tổng hợp cử tri', 'Tổng hợp', ...)
    2. Total/aggregate sheets ('Tổng', 'SHEET TỔNG', 'TH', 'biểu', 'Khu Phố', ...)
    3. All other sheets (to handle any remaining edge cases)

    Each group keeps the workbook's original sheet order.
    """
    group1, group2, group3 = [], [], []
    for name in wb.sheetnames:
        n = _norm(name)
        # Check group 1: 'tong hop' substring (handles slight variations)
        if n in _SUMMARY_SHEET_NAMES or "tong hop" in n:
            group1.append(wb[name])
        # Check group 2: exact match against known total sheet names
        elif n in _TOTAL_SHEET_NAMES:
            group2.append(wb[name])
        else:
            group3.append(wb[name])
    return group1 + group2 + group3


def extract_from_text_cell(cell_value) -> dict | None:
    """
    Try to extract (tong, nam, nu) from a free-text string cell.
    Handles patterns like:
      'Tổng số cử tri của khu vực bỏ phiếu là:1266 người; trong đó có: 607 Nam; 659 Nữ.'
      'Tổng số cử tri ... là 1150 người; trong đó có: 554 Nam, 596 Nữ.'
    """
    if not isinstance(cell_value, str):
        return None
    text = cell_value.strip()
    if not text:
        return None

    # Try primary pattern
    m = _TEXT_TOTAL_RE.search(text)
    if m:
        tong, nam, nu = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if tong > 0 and (nam + nu == tong or nam > 0 and nu > 0):
            return {"tong": tong, "nam": nam, "nu": nu}

    # Try alt pattern 1: X Nam, Y Nữ
    m = _TEXT_ALT_RE.search(text)
    if m:
        tong, nam, nu = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if tong > 0 and nam > 0 and nu > 0:
            return {"tong": tong, "nam": nam, "nu": nu}

    # Try alt pattern 2: Nam: X, Nữ: Y
    m = _TEXT_NAM_NU_RE.search(text)
    if m:
        tong, nam, nu = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if tong > 0 and nam > 0 and nu > 0:
            return {"tong": tong, "nam": nam, "nu": nu}

    return None


def find_total_row(ws, wb_formulas=None) -> dict | None:
    """
    Scan worksheet for Tổng/Nam/Nữ data using multiple strategies:

    Strategy 0 (FORMULA): When wb_formulas provided, scan formula cells that
      return None in data_only mode. Parse &CellRef& patterns from the formula
      string and read those numeric cells directly.

    Strategy 1 (TEXT): Scan cell values for free-text sentence pattern.

    Strategy 2 (STRUCTURED): 'Tổng' label row with tong == nam + nu.

    Strategy 3 (SMART TRIPLET): Any 3 numbers in close proximity (same row or
      adjacent rows) where tong = nam + nu and tong > 50.

    Strategy 4 (FALLBACK): 'Tổng' label row with any 3 positive nums.
    """
    all_rows = list(ws.iter_rows(values_only=True))

    # ── Strategy 0: Parse formula cells via wb_formulas ──────────────────────
    if wb_formulas and ws.title in wb_formulas.sheetnames:
        ws_f = wb_formulas[ws.title]
        for row_f in ws_f.iter_rows():
            for cell_f in row_f:
                val = cell_f.value
                # Only care about formula cells that look like text concatenations
                if not isinstance(val, str) or not val.startswith('='):
                    continue
                # Check if the formula text contains tong/nam/nu keywords
                norm_formula = normalize_text(val)
                if 'nguoi' not in norm_formula and 'ng' not in norm_formula:
                    continue
                # Extract cell references like &B17& &C17& &D17&
                refs = _FORMULA_REF_RE.findall(val)
                nums_from_refs = []
                for ref in refs:
                    try:
                        ref_cell = ws_f[ref]  # same sheet
                        ref_val = extract_numeric(ref_cell.value)
                        if ref_val is not None:
                            nums_from_refs.append(ref_val)
                    except Exception:
                        pass
                if len(nums_from_refs) >= 3:
                    tong, nam, nu = nums_from_refs[0], nums_from_refs[1], nums_from_refs[2]
                    if tong == nam + nu and tong > 0:
                        return {"tong": tong, "nam": nam, "nu": nu}
                    # Try all triples
                    for i in range(len(nums_from_refs) - 2):
                        t, m, f = nums_from_refs[i], nums_from_refs[i+1], nums_from_refs[i+2]
                        if t == m + f and t > 0:
                            return {"tong": t, "nam": m, "nu": f}

    # ── Strategy 1: Text sentence extraction ─────────────────────────────
    for row in all_rows:
        for cell in row:
            result = extract_from_text_cell(cell)
            if result:
                return result

    # ── Strategy 2: Structured table row with Tổng label ───────────────────
    best_candidate = None

    for row in all_rows:
        label_found = any(is_total_row_label(cell) for cell in row)
        if not label_found:
            continue

        nums = [extract_numeric(c) for c in row if extract_numeric(c) is not None]

        if len(nums) >= 3:
            tong, nam, nu = nums[0], nums[1], nums[2]
            if tong == nam + nu:
                return {"tong": tong, "nam": nam, "nu": nu}
            for i in range(len(nums) - 2):
                t, m, f = nums[i], nums[i+1], nums[i+2]
                if t == m + f and t > 0:
                    return {"tong": t, "nam": m, "nu": f}
            if tong > 0 and nam > 0 and nu > 0:
                if best_candidate is None or tong > best_candidate["tong"]:
                    best_candidate = {"tong": tong, "nam": nam, "nu": nu}

        elif len(nums) == 1:
            row_idx = all_rows.index(row)
            if row_idx + 1 < len(all_rows):
                next_nums = [extract_numeric(c) for c in all_rows[row_idx + 1]
                             if extract_numeric(c) is not None]
                if len(next_nums) >= 2:
                    return {"tong": nums[0], "nam": next_nums[0], "nu": next_nums[1]}

    if best_candidate:
        return best_candidate

    # ── Strategy 3: Smart triplet scan (tong = nam + nu, tong > 50) ─────────
    # Scan the LAST 100 rows for any row with exactly 2-4 numbers where t=m+f
    for row in reversed(all_rows[-200:] if len(all_rows) > 200 else all_rows):
        nums = [extract_numeric(c) for c in row if extract_numeric(c) is not None]
        if len(nums) < 2:
            continue
        for i in range(len(nums) - 1):
            m_val, f_val = nums[i], nums[i+1]
            if m_val > 0 and f_val > 0:
                t_val = m_val + f_val
                if t_val > 50:  # Reasonable voter count threshold
                    # Check if t_val appears in nearby rows too
                    # Look backwards for t_val
                    idx = all_rows.index(row) if row in all_rows else -1
                    if idx > 0:
                        prev_row_nums = [extract_numeric(c) for c in all_rows[idx-1]
                                         if extract_numeric(c) is not None]
                        if t_val in prev_row_nums:
                            return {"tong": t_val, "nam": m_val, "nu": f_val}

    # ── Strategy 4: Ultimate fallback ─────────────────────────────────
    for row in reversed(all_rows):
        nums = [extract_numeric(c) for c in row if extract_numeric(c) is not None]
        if len(nums) >= 3 and any(is_total_row_label(c) for c in row):
            return {"tong": nums[0], "nam": nums[1], "nu": nums[2]}

    return None


_STOP_KEYWORDS = re.compile(
    r"tong\s*so\s*cu\s*tri|cu\s*tri\s*tham\s*gia|danh\s*sach\s*duoc\s*lap|nguoi\s*lap\s*bieu",
    re.IGNORECASE | re.UNICODE,
)

def _parse_dob(cell_val) -> datetime.date | None:
    """Parse a date-of-birth cell (datetime, date, or string) → date or None."""
    if isinstance(cell_val, datetime.datetime):
        return cell_val.date()
    if isinstance(cell_val, datetime.date):
        return cell_val
    if isinstance(cell_val, str):
        s = cell_val.strip()
        for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y", "%d.%m.%Y"):
            try:
                return datetime.datetime.strptime(s, fmt).date()
            except ValueError:
                continue
    return None


def _is_x(val) -> bool:
    """True if cell contains 'x' or 'X' (ignores spaces, dots, dashes... around 'x')."""
    if val is None:
        return False
    return "x" in str(val).lower()


def count_voter_stats(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict:
    """
    Scan ONE voter-list sheet (Tổng or Tổng Hợp) and return ALL stats in one pass.

    Column layout (1-based):
      A(1) = STT       B(2) = Họ tên   C(3) = Ngày sinh
      D(4) = Nam (x)   E(5) = Nữ  (x)
      K(11) = Bầu ĐBQH (x)
      L(12) = Bầu HĐND tỉnh (x)
      M(13) = Bầu HĐND phường/xã (x)

    Returns:
      tong, nam, nu       → for columns F, G, H
      ct18, elderly       → for columns K, L in summary
      qh_total/nam/nu     → for columns M, N, O
      tinh_total/nam/nu   → for columns P, Q, R
      xa_total/nam/nu     → for columns S, T, U
    """
    stats = dict(
        tong=0,  nam=0,  nu=0,
        ct18=0,  elderly=0,
        qh_total=0,   qh_nam=0,   qh_nu=0,
        tinh_total=0, tinh_nam=0, tinh_nu=0,
        xa_total=0,   xa_nam=0,   xa_nu=0,
    )

    DOB_18_FROM = datetime.date(2007, 3, 16)
    DOB_18_TO   = datetime.date(2008, 3, 15)
    DOB_80_CUT  = datetime.date(1946, 3, 16)

    # ── Find data start: first row where col C has a valid DOB ───────────────
    data_start = 19
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=50, values_only=True)):
        if _parse_dob(row[2] if len(row) > 2 else None):
            data_start = idx + 1
            break

    # ── Single pass over all voter rows ─────────────────────────────────────
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        # Stop at footer (e.g. "Tổng số cử tri", "Người lập biểu")
        col_b = normalize_text(str(row[1] if len(row) > 1 else ""))
        if _STOP_KEYWORDS.search(col_b):
            break

        # We shouldn't skip the row purely because DOB parsing fails. 
        # Sometimes DOB is malformed (e.g. "1975", "28//2004"). 
        # Check if they have a name in col B OR some raw text in col C
        has_name = bool(str(row[1]).strip()) if len(row) > 1 and row[1] else False
        has_raw_dob = bool(str(row[2]).strip()) if len(row) > 2 and row[2] else False
        
        dob = _parse_dob(row[2] if len(row) > 2 else None)
        
        if not has_name and not has_raw_dob:
            continue   # not a voter row (both name and DOB are completely empty)

        co_nam = _is_x(row[3] if len(row) > 3 else None)   # D = Nam
        co_nu  = _is_x(row[4] if len(row) > 4 else None)   # E = Nữ

        # ── F, G, H: total count ────────────────────────────────────────────
        stats["tong"] += 1
        if co_nam: stats["nam"] += 1
        if co_nu:  stats["nu"]  += 1

        # ── Age groups (K, L in summary) ────────────────────────────────────
        if dob:
            if DOB_18_FROM <= dob <= DOB_18_TO:
                stats["ct18"] += 1
            if dob < DOB_80_CUT:
                stats["elderly"] += 1

        # ── Election marks (M-U in summary) ─────────────────────────────────
        co_qh   = _is_x(row[10] if len(row) > 10 else None)  # K = QH
        co_tinh = _is_x(row[11] if len(row) > 11 else None)  # L = Tinh
        co_xa   = _is_x(row[12] if len(row) > 12 else None)  # M = Phường/Xã

        if co_qh:
            stats["qh_total"] += 1
            if co_nam: stats["qh_nam"] += 1
            if co_nu:  stats["qh_nu"]  += 1
        if co_tinh:
            stats["tinh_total"] += 1
            if co_nam: stats["tinh_nam"] += 1
            if co_nu:  stats["tinh_nu"]  += 1
        if co_xa:
            stats["xa_total"] += 1
            if co_nam: stats["xa_nam"] += 1
            if co_nu:  stats["xa_nu"]  += 1

    return stats


def _has_voter_headers(ws: openpyxl.worksheet.worksheet.Worksheet) -> bool:
    """Kiểm tra xem sheet có chứa các cột tiêu chuẩn của danh sách cử tri không."""
    reqs = ["stt", "ho va ten", "ngay", "nam", "nu", "can cuoc"]
    for row in ws.iter_rows(min_row=1, max_row=25, values_only=True):
        row_str = " ".join([normalize_text(str(c)) for c in row if c])
        # Cần ít nhất 4 từ khóa để xác nhận
        if sum(1 for req in reqs if req in row_str) >= 4:
            return True
    return False

def _find_voter_list_sheet(wb: openpyxl.Workbook):
    """
    Find the best voter-list sheet to scan for 'x' marks.
    Priority order:
      1. Sheet named 'Tổng Hợp' / 'Tổng hợp cử tri'
      2. Sheet named 'Tổng' / 'Biểu tổng'
    MUST strictly contain actual voter table headers.
    """
    named_tong_hop, named_tong, others = [], [], []
    for sheetname in wb.sheetnames:
        n = normalize_text(sheetname)
        # Bỏ qua các sheet Tổ Cụ Thể (Tổ 1, Tổ 2...)
        if "to " in n and "tong" not in n:
            continue
            
        if "tong hop" in n:
            named_tong_hop.append(sheetname)
        elif "tong" in n:
            named_tong.append(sheetname)
        else:
            others.append(sheetname)

    # Chỉ quét các sheet Tổng
    for sheetname in named_tong_hop + named_tong:
        ws = wb[sheetname]
        if _has_voter_headers(ws):
            for row in ws.iter_rows(min_row=1, max_row=50, values_only=True):
                if _parse_dob(row[2] if len(row) > 2 else None):
                    return ws

    # Fallback cho các file không ghi rõ chữ Tổng nhưng có bảng
    for sheetname in others:
        ws = wb[sheetname]
        if _has_voter_headers(ws):
            for row in ws.iter_rows(min_row=1, max_row=50, values_only=True):
                if _parse_dob(row[2] if len(row) > 2 else None):
                    return ws
    
    return None


@st.cache_data(show_spinner=False, ttl=3600*24)
def _process_source_bytes(file_bytes: bytes) -> dict:
    """Core logic extracted to allow robust caching based on file contents."""
    result = {
        "tong": None, "nam": None, "nu": None,
        "ct18": None, "elderly": None,
        "qh_total": None, "qh_nam": None, "qh_nu": None,
        "tinh_total": None, "tinh_nam": None, "tinh_nu": None,
        "xa_total": None, "xa_nam": None, "xa_nu": None,
        "error": None,
    }
    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        result["error"] = f"Không mở được file: {e}"
        return result

    voter_ws = _find_voter_list_sheet(wb)
    if voter_ws is None:
        result["error"] = "Không tìm thấy sheet có dữ liệu cử tri (ngày sinh)"
        return result

    try:
        stats = count_voter_stats(voter_ws)
        if stats["tong"] > 0:
            result.update(stats)
        else:
            result["error"] = f"Sheet '{voter_ws.title}': không đếm được cử tri (0 hàng)"
    except Exception as e:
        result["error"] = f"Lỗi khi đọc sheet '{voter_ws.title}': {e}"

    return result

def process_source_file(filepath: str) -> dict:
    """
    Read a source .xlsx file and return all extracted voter statistics.
    Converts file to bytes first to utilize Streamlit's robust content-based caching.
    """
    try:
        with open(filepath, "rb") as f:
            file_bytes = f.read()
        return _process_source_bytes(file_bytes)
    except Exception as e:
        return {"error": f"Lỗi đọc file từ đĩa: {e}"}


# ─────────────────────────────────────────────────────────────────────────────
# Summary file update
# ─────────────────────────────────────────────────────────────────────────────

def col_letter_to_idx(letter: str) -> int:
    """Convert column letter like 'A' → 1, 'F' → 6."""
    result = 0
    for c in letter.upper():
        result = result * 26 + (ord(c) - ord('A') + 1)
    return result


def find_name_column(ws) -> int | None:
    """
    Scan rows to find the column that contains khu phố names.
    Returns 1-based column index or None.
    """
    for row in ws.iter_rows(min_row=1, max_row=30, values_only=True):
        for col_idx, cell in enumerate(row, start=1):
            if cell and isinstance(cell, str) and len(cell.strip()) > 3:
                norm = normalize_text(cell)
                # Heuristic: column contains something like "Khu phố 1", "Thôn", etc.
                if any(k in norm for k in ["khu pho", "thon", "ban ", "xa ", "phuong"]):
                    return col_idx
    return None


_RE_LEADING_NUM = re.compile(r'^\d+[\.\-\s]+')
_RE_PAREN       = re.compile(r'\(.*?\)')    # strip (đã in), (đã nộp), ...

def _strip_file_prefix(name: str) -> str:
    """
    Normalise a file-stem or row-cell name for matching:
    1. Strip leading number: '26. Khu phố ...'  → 'Khu phố ...'
    2. Strip parenthetical suffixes: '... (đã in)' → '...'
    """
    name = _RE_LEADING_NUM.sub('', name).strip()
    name = _RE_PAREN.sub('', name).strip()
    return name


def _name_key(name: str) -> str:
    """Full normalisation: strip prefix/suffix + remove diacritics + lowercase."""
    return normalize_text(_strip_file_prefix(name))


def _token_overlap(a: str, b: str) -> float:
    """
    Return the Jaccard-like token overlap between two normalised strings.
    Return 0.0 aggressively if trailing identifiers mismatch (e.g. '1' vs '2')
    or if core contrasting words exist (e.g. 'Tien' vs 'Hai').
    """
    stop = {'khu', 'pho', 'to', 'thon', 'ban', 'đon', 'vi', 'bau', 'cu', 'so', 'phuong', 'xa', 'ap', 'tô'}
    
    tokens_a = a.split()
    tokens_b = b.split()
    
    def get_id_token(tokens):
        if not tokens: return None
        if any(c.isdigit() for c in tokens[-1]): return tokens[-1]
        if len(tokens) > 1 and any(c.isdigit() for c in tokens[-2]) and tokens[-1] == '.xlsx': return tokens[-2]
        return None

    num_a = get_id_token(tokens_a)
    num_b = get_id_token(tokens_b)
    
    if num_a and num_b and num_a != num_b:
        return 0.0

    core_a = {t for t in tokens_a if t not in stop and not any(c.isdigit() for c in t) and not t.endswith('.xlsx')}
    core_b = {t for t in tokens_b if t not in stop and not any(c.isdigit() for c in t) and not t.endswith('.xlsx')}
    
    if core_a and core_b:
        if (core_a - core_b) and (core_b - core_a):
            # Contradicting distinct names detected
            return 0.0

    ta = {t for t in tokens_a if len(t) > 1 or t.isdigit()}
    tb = {t for t in tokens_b if len(t) > 1 or t.isdigit()}
    
    if not ta or not tb:
        return 0.0
        
    return len(ta & tb) / max(len(ta), len(tb))


def _fuzzy_match(norm_row: str, lookup: dict, threshold: float = 0.6):
    """
    Fuzzy fallback: find the lookup key with the highest token overlap
    with norm_row, only if overlap >= threshold.
    Returns the matched (orig_key, data) tuple or None.
    """
    best_score, best_entry = 0.0, None
    for lk, val in lookup.items():
        # Fast short-circuit: substring match
        if lk and (lk in norm_row or norm_row in lk):
            return val
        score = _token_overlap(norm_row, lk)
        if score > best_score:
            best_score, best_entry = score, val
    if best_score >= threshold:
        return best_entry
    return None


def update_summary_file(summary_path: str, data_map: dict, log_fn=None) -> bytes:
    """
    Update summary xlsx with aggregated data.

    data_map: {display_name: {'tong':..,'nam':..,'nu':..,'ct18':..,'elderly':..}}
      - Keys are the original source file stems (e.g. '31. Khu phố Lê Lợi').

    Column mapping (user-confirmed):
      F (6)  = Tổng số cử tri
      G (7)  = Nam
      H (8)  = Nữ
      K (11) = Cử tri 18 tuổi lần đầu bỏ phiếu
      L (12) = Cử tri cao tuổi (> 80)

    Name matching: column B (col 2) holds the unit names to match against.
    Both the row value and the file stem are stripped of leading "N. " prefixes,
    then compared with normalize_text() for diacritic-insensitive matching.

    Returns the updated workbook as bytes.
    """
    wb = load_workbook(summary_path)
    ws = wb.active

    # ── Column indices (1-based, user-confirmed) ────────────────────────────
    COL_NAME  = 3   # C = Tổ thôn, bản, khu phố (tên khu phố khớp với tên file)
    COL_F     = 6   # Tổng số cử tri
    COL_G     = 7   # Nam
    COL_H     = 8   # Nữ
    COL_K     = 11  # Cử tri 18 tuổi lần đầu bỏ phiếu
    COL_L     = 12  # Cử tri cao tuổi (> 80 tuổi)
    COL_M     = 13  # Bầu ĐBQH - Tổng số
    COL_N     = 14  # Bầu ĐBQH - Nam
    COL_O     = 15  # Bầu ĐBQH - Nữ
    COL_P     = 16  # Bầu ĐBHĐND tỉnh - Tổng số
    COL_Q     = 17  # Bầu ĐBHĐND tỉnh - Nam
    COL_R     = 18  # Bầu ĐBHĐND tỉnh - Nữ
    COL_S     = 19  # Bầu ĐBHĐND cấp xã - Tổng số
    COL_T     = 20  # Bầu ĐBHĐND cấp xã - Nam
    COL_U     = 21  # Bầu ĐBHĐND cấp xã - Nữ

    # Map from data key → column index (for clearing + writing)
    DATA_COLS = [
        ("tong",       COL_F),
        ("nam",        COL_G),
        ("nu",         COL_H),
        ("ct18",       COL_K),
        ("elderly",    COL_L),
        ("qh_total",   COL_M),
        ("qh_nam",     COL_N),
        ("qh_nu",      COL_O),
        ("tinh_total", COL_P),
        ("tinh_nam",   COL_Q),
        ("tinh_nu",    COL_R),
        ("xa_total",   COL_S),
        ("xa_nam",     COL_T),
        ("xa_nu",      COL_U),
    ]

    # ── Build normalised lookup from data_map ────────────────────────────────
    # Both the full key and the cleaned key are stored so either can match a row.
    lookup: dict[str, tuple] = {}
    for orig_key, data in data_map.items():
        lookup[_name_key(orig_key)]          = (orig_key, data)
        lookup[normalize_text(orig_key)]     = (orig_key, data)   # keep raw norm too

    updated_rows = 0
    for row in ws.iter_rows():
        if len(row) < COL_NAME:
            continue
        name_cell = row[COL_NAME - 1]   # 0-indexed → column C
        if not name_cell.value:
            continue

        raw_name  = str(name_cell.value).strip()
        norm_name = _name_key(raw_name)

        # 1) Exact match (normalised)
        entry = lookup.get(norm_name)
        # 2) Fuzzy match: substring + token overlap ≥ 60%
        if entry is None:
            entry = _fuzzy_match(norm_name, lookup, threshold=0.6)

        if entry is None:
            continue

        _, data = entry
        row_idx  = name_cell.row

        # Clear existing values and write new ones
        for key, col in DATA_COLS:
            cell = ws.cell(row=row_idx, column=col)
            cell.value = None                    # clear first
            v = data.get(key)
            if v is not None:
                cell.value = v

        updated_rows += 1
        if log_fn:
            log_fn(f"  ✅ Cập nhật hàng '{raw_name}': "
                   f"Tổng={data.get('tong')}, Nam={data.get('nam')}, "
                   f"Nữ={data.get('nu')}, CT18={data.get('ct18')}, "
                   f"Cao tuổi={data.get('elderly')}, "
                   f"QH={data.get('qh_total')}, "
                   f"HĐND tỉnh={data.get('tinh_total')}, "
                   f"HĐND xã={data.get('xa_total')}")

    if log_fn:
        log_fn(f"\n📊 Tổng số hàng đã cập nhật: **{updated_rows}**")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────────────────────────────────────
# Source file collection (local or GDrive)
# ─────────────────────────────────────────────────────────────────────────────

def collect_source_files(source_input: str, tmp_dir: str) -> list:
    """
    Returns list of (display_name, filepath) tuples.
    Handles:
      - local folder path
      - Google Drive folder URL
    De-duplicates by filename (keep random one).
    """
    files_raw = []

    if "drive.google.com" in source_input or "docs.google.com" in source_input:
        folder_id = parse_gdrive_folder_id(source_input)
        if not folder_id:
            raise ValueError("Không nhận ra folder ID từ link Google Drive.")
        dl_dir = os.path.join(tmp_dir, "gdrive_src")
        os.makedirs(dl_dir, exist_ok=True)
        files_raw = download_gdrive_folder(folder_id, dl_dir)
    else:
        # Local path
        path = source_input.strip().strip('"').strip("'")
        if not os.path.isdir(path):
            raise ValueError(f"Không tìm thấy folder: {path}")
        for fn in os.listdir(path):
            if fn.lower().endswith(".xlsx") and not fn.startswith("~"):
                files_raw.append(os.path.join(path, fn))

    # De-duplicate by filename stem
    name_map: dict[str, list] = {}
    for fp in files_raw:
        stem = os.path.splitext(os.path.basename(fp))[0]
        name_map.setdefault(stem, []).append(fp)

    result = []
    for stem, paths in name_map.items():
        chosen = random.choice(paths)
        result.append((stem, chosen))

    return sorted(result, key=lambda x: x[0])


def get_summary_file_bytes(summary_input: str, tmp_dir: str) -> str:
    """Download or resolve the summary file. Returns local filepath."""
    if "drive.google.com" in summary_input or "docs.google.com" in summary_input:
        file_id = parse_gdrive_file_id(summary_input)
        if not file_id:
            raise ValueError("Không nhận ra file ID từ link Google Drive.")
        dest = os.path.join(tmp_dir, "summary_file.xlsx")
        download_gdrive_file(file_id, dest)
        return dest
    else:
        path = summary_input.strip().strip('"').strip("'")
        if not os.path.isfile(path):
            raise ValueError(f"Không tìm thấy file: {path}")
        return path


# ─────────────────────────────────────────────────────────────────────────────
# Streamlit UI
# ─────────────────────────────────────────────────────────────────────────────

def main():
    st.set_page_config(
        page_title="Tổng hợp Dữ liệu Cử tri",
        page_icon="🗳️",
        layout="wide",
        initial_sidebar_state="expanded",
    )

    # ── Custom CSS ──────────────────────────────────────────────────────────
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');

    html, body, [class*="css"] {
        font-family: 'Inter', sans-serif;
    }

    .main-header {
        background: linear-gradient(135deg, #1a237e 0%, #283593 50%, #3949ab 100%);
        padding: 2rem 2.5rem;
        border-radius: 16px;
        margin-bottom: 2rem;
        box-shadow: 0 4px 24px rgba(26,35,126,0.2);
    }
    .main-header h1 {
        color: white;
        font-size: 2rem;
        font-weight: 700;
        margin: 0;
        letter-spacing: -0.5px;
    }
    .main-header p {
        color: rgba(255,255,255,0.8);
        margin: 0.4rem 0 0 0;
        font-size: 1rem;
    }

    .info-card {
        background: linear-gradient(135deg, #e3f2fd, #bbdefb);
        border-left: 4px solid #1976d2;
        border-radius: 10px;
        padding: 1rem 1.25rem;
        margin-bottom: 1rem;
    }
    .info-card h4 { color: #0d47a1; margin: 0 0 0.4rem 0; }
    .info-card p  { color: #1565c0; margin: 0; font-size: 0.88rem; }

    .log-box {
        background: #0d1117;
        color: #c9d1d9;
        border-radius: 10px;
        padding: 1.25rem;
        font-family: 'Courier New', monospace;
        font-size: 0.85rem;
        line-height: 1.7;
        max-height: 400px;
        overflow-y: auto;
        border: 1px solid #30363d;
    }

    .metric-card {
        background: white;
        border-radius: 12px;
        padding: 1.25rem;
        text-align: center;
        box-shadow: 0 2px 12px rgba(0,0,0,0.08);
        border: 1px solid #e0e0e0;
    }
    .metric-card .value {
        font-size: 2.2rem;
        font-weight: 700;
        color: #1a237e;
    }
    .metric-card .label {
        color: #666;
        font-size: 0.85rem;
        margin-top: 0.2rem;
    }

    div[data-testid="stButton"] > button {
        background: linear-gradient(135deg, #1a237e, #3949ab);
        color: white;
        border: none;
        border-radius: 10px;
        padding: 0.65rem 2rem;
        font-weight: 600;
        font-size: 1rem;
        width: 100%;
        transition: all 0.2s ease;
        box-shadow: 0 4px 14px rgba(57,73,171,0.35);
    }
    div[data-testid="stButton"] > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 6px 20px rgba(57,73,171,0.5);
    }

    .stDownloadButton > button {
        background: linear-gradient(135deg, #2e7d32, #43a047) !important;
        color: white !important;
        border: none !important;
        border-radius: 10px !important;
        width: 100% !important;
        font-weight: 600 !important;
    }

    section[data-testid="stSidebar"] {
        background: linear-gradient(180deg, #f8f9ff 0%, #eff1ff 100%);
    }

    /* Compact buttons in sidebar */
    div[data-testid="stSidebar"] div[data-testid="stButton"] > button {
        height: auto;
        padding-top: 0.4rem;
        padding-bottom: 0.4rem;
        font-size: 0.85rem;
    }
    </style>
    """, unsafe_allow_html=True)

    # ── Header ──────────────────────────────────────────────────────────────
    st.markdown("""
    <div class="main-header">
        <h1>🗳️ Tổng hợp Dữ liệu Cử tri</h1>
        <p>Tự động đọc file Excel nguồn từ nhiều khu phố/thôn và cập nhật vào bảng tổng hợp</p>
    </div>
    """, unsafe_allow_html=True)

    # ── Sidebar inputs ───────────────────────────────────────────────────────
    with st.sidebar:
        st.markdown("## ⚙️ Cấu hình")
        st.markdown("---")

        # ── Input mode toggle
        input_mode = st.radio(
            "📥 Chế độ nhập liệu",
            ["Upload file lên trực tiếp ▲", "Nhập đường dẫn / Google Drive"],
            help="Chọn cách cung cấp dữ liệu"
        )
        st.markdown("---")

        uploaded_sources = None
        uploaded_summary = None
        source_input = ""
        summary_input = ""

        if input_mode == "Upload file lên trực tiếp ▲":
            # Initialize keys for resetting uploaders
            if "source_key" not in st.session_state:
                st.session_state.source_key = 0
            if "summary_key" not in st.session_state:
                st.session_state.summary_key = 0

            st.markdown("### 📁 Nguồn dữ liệu")
            uploaded_sources = st.file_uploader(
                "Chọn nhiều file .xlsx nguồn (mỗi file = 1 khu phố)",
                type=["xlsx"],
                accept_multiple_files=True,
                key=f"source_uploader_{st.session_state.source_key}",
                help="Chọn tất cả file .xlsx của các khu phố cùng lúc"
            )
            if uploaded_sources:
                st.success(f"✅ Đã chọn {len(uploaded_sources)} file")
                if st.button("🗑️ Xoá tất cả file nguồn", key="clear_source"):
                    st.session_state.source_key += 1
                    st.rerun()

            st.markdown("### 📊 File tổng hợp")
            uploaded_summary = st.file_uploader(
                "Chọn file BIỂU TỔNG HỢP.xlsx",
                type=["xlsx"],
                accept_multiple_files=False,
                key=f"summary_uploader_{st.session_state.summary_key}",
                help="File này sẽ được cập nhật cột F, G, H, K, L"
            )
            if uploaded_summary:
                st.success(f"✅ Đã chọn: {uploaded_summary.name}")
                if st.button("🗑️ Xoá file tổng hợp", key="clear_summary"):
                    st.session_state.summary_key += 1
                    st.rerun()

        else:
            st.markdown("### 📁 Nguồn dữ liệu")
            source_input = st.text_area(
                "Folder chứa file .xlsx nguồn",
                placeholder=(
                    "Đường dẫn local:\n  /Users/ten/Documents/du_lieu\n\n"
                    "Google Drive:\n  https://drive.google.com/drive/folders/..."
                ),
                height=100,
                help="Mỗi file .xlsx = một khu phố",
            )
            st.markdown("### 📊 File tổng hợp")
            summary_input = st.text_area(
                "File BIỂU TỔNG HỢP.xlsx",
                placeholder=(
                    "Đường dẫn local:\n  /Users/ten/Documents/BIEU_TONG_HOP.xlsx\n\n"
                    "Google Drive:\n  https://drive.google.com/file/d/..."
                ),
                height=100,
                help="File này sẽ được cập nhật cột F, G, H, K, L"
            )

        st.markdown("---")

        with st.expander("📋 Cột sẽ được cập nhật"):
            st.markdown("""
            | Cột | Nội dung |
            |-----|----------|
            | **F** | Tổng số cử tri |
            | **G** | Nam |
            | **H** | Nữ |
            | **K** | Cử tri 18 tuổi lần đầu |
            | **L** | Cử tri cao tuổi (>80 tuổi) |
            """)

        with st.expander("ℹ️ Ghi chú"):
            st.markdown("""
            - **Ngày bầu cử:** 15/3/2026  
            - **CT 18 tuổi:** sinh từ 16/3/2007 → 15/3/2008  
            - **Cao tuổi >80:** sinh trước/bằng 15/3/1946  
            - File trùng tên → chọn ngẫu nhiên 1 file  
            - Google Drive folder phải **public** (Anyone with the link)
            """)

        st.markdown("---")
        process_btn = st.button("▶ Bắt đầu xử lý", use_container_width=True)

    # ── Main area ────────────────────────────────────────────────────────────
    if not process_btn:
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("""
            <div class="info-card">
                <h4>📁 Bước 1 — Nhập đường dẫn folder nguồn</h4>
                <p>Folder chứa nhiều file .xlsx, mỗi file là một khu phố/thôn/bản.
                Hỗ trợ đường dẫn local và link Google Drive folder (public).</p>
            </div>
            """, unsafe_allow_html=True)
        with col2:
            st.markdown("""
            <div class="info-card">
                <h4>📊 Bước 2 — Nhập đường dẫn file tổng hợp</h4>
                <p>File BIỂU TỔNG HỢP DANH SÁCH CỬ TRI.xlsx sẽ được đọc và
                cập nhật tự động. File gốc không bị thay đổi — bạn tải bản mới về.</p>
            </div>
            """, unsafe_allow_html=True)

        st.markdown("""
        <div class="info-card">
            <h4>🔄 Bước 3 — Bấm "Bắt đầu xử lý"</h4>
            <p>Hệ thống sẽ tự động đọc từng file nguồn, trích xuất Tổng/Nam/Nữ
            và đếm cử tri theo nhóm tuổi, sau đó cập nhật vào bảng tổng hợp.
            Preview kết quả và tải file .xlsx mới về máy.</p>
        </div>
        """, unsafe_allow_html=True)
        return

    # ── Processing ────────────────────────────────────────────────────────────
    is_upload_mode = (input_mode == "Upload file lên trực tiếp ▲")

    if is_upload_mode:
        if not uploaded_sources:
            st.error("⚠️ Vui lòng upload ít nhất 1 file .xlsx nguồn.")
            return
        if uploaded_summary is None:
            st.error("⚠️ Vui lòng upload file BIỂU TỔNG HỢP.xlsx.")
            return
    else:
        if not source_input.strip():
            st.error("⚠️ Vui lòng nhập đường dẫn folder nguồn.")
            return
        if not summary_input.strip():
            st.error("⚠️ Vui lòng nhập đường dẫn file tổng hợp.")
            return

    log_lines: list[str] = []
    progress_placeholder = st.empty()
    log_placeholder = st.empty()

    def log(msg: str):
        log_lines.append(msg)
        display = "\n".join(log_lines)
        log_placeholder.markdown(
            f'<div class="log-box"><pre style="margin:0;white-space:pre-wrap;">{display}</pre></div>',
            unsafe_allow_html=True
        )

    with tempfile.TemporaryDirectory() as tmp_dir:
        log("🚀 Bắt đầu xử lý...\n")

        # Step 1: collect source files
        if is_upload_mode:
            # Save uploaded files to temp dir
            source_files = []
            for f in uploaded_sources:
                stem = os.path.splitext(f.name)[0]
                dst = os.path.join(tmp_dir, f.name)
                with open(dst, 'wb') as out:
                    out.write(f.read())
                source_files.append((stem, dst))
            source_files = sorted(source_files, key=lambda x: x[0])
            log(f"   → Đã nhận {len(source_files)} file .xlsx từ upload\n")
        else:
            try:
                log("📂 Đang đọc folder nguồn...")
                source_files = collect_source_files(source_input, tmp_dir)
                log(f"   → Tìm thấy {len(source_files)} file .xlsx\n")
            except Exception as e:
                st.error(f"❌ Lỗi đọc folder nguồn: {e}")
                log(f"❌ Lỗi: {e}")
                return

        if not source_files:
            st.warning("⚠️ Không tìm thấy file .xlsx nào trong folder nguồn.")
            return

        # Step 2: load summary file
        if is_upload_mode:
            summary_path = os.path.join(tmp_dir, uploaded_summary.name)
            with open(summary_path, 'wb') as out:
                out.write(uploaded_summary.read())
            log("   → Đã nhận file tổng hợp từ upload\n")
        else:
            try:
                log("📊 Đang tải file tổng hợp...")
                summary_path = get_summary_file_bytes(summary_input, tmp_dir)
                log("   → Tải thành công\n")
            except Exception as e:
                st.error(f"❌ Lỗi đọc file tổng hợp: {e}")
                log(f"❌ Lỗi: {e}")
                return

        # Step 3: process each source file
        data_map: dict = {}
        errors: list = []

        progress_bar = progress_placeholder.progress(0, text="Đang xử lý file nguồn...")
        n = len(source_files)

        for i, (stem, filepath) in enumerate(source_files):
            log(f"📄 [{i+1}/{n}] Đang xử lý: {stem} ...")
            result = process_source_file(filepath)

            if result["error"]:
                errors.append((stem, result["error"]))
                log(f"   ⚠️  Lỗi: {result['error']}")
            else:
                key = normalize_text(stem)
                data_map[key] = result
                log(f"   ✅ Tổng={result['tong']}, Nam={result['nam']}, "
                    f"Nữ={result['nu']}, CT18={result['ct18']}, "
                    f"Cao tuổi={result['elderly']}")

            progress_bar.progress((i + 1) / n, text=f"Đang xử lý [{i+1}/{n}]: {stem}")

        progress_bar.progress(1.0, text="Hoàn thành đọc file nguồn ✓")

        # Step 4: update summary file
        log("\n📝 Đang ghi vào file tổng hợp...")
        try:
            updated_bytes = update_summary_file(summary_path, data_map, log_fn=log)
        except Exception as e:
            st.error(f"❌ Lỗi cập nhật file tổng hợp: {e}")
            log(f"❌ Lỗi: {e}")
            return

        log("\n✅ Hoàn thành! File sẵn sàng để tải về.")

        # ── Summary metrics ──────────────────────────────────────────────────
        st.markdown("<br>", unsafe_allow_html=True)
        col1, col2, col3 = st.columns(3)
        with col1:
            st.markdown(f"""
            <div class="metric-card">
                <div class="value">{len(source_files)}</div>
                <div class="label">File nguồn đã đọc</div>
            </div>""", unsafe_allow_html=True)
        with col2:
            st.markdown(f"""
            <div class="metric-card">
                <div class="value">{len(data_map)}</div>
                <div class="label">Khu phố cập nhật thành công</div>
            </div>""", unsafe_allow_html=True)
        with col3:
            st.markdown(f"""
            <div class="metric-card">
                <div class="value">{len(errors)}</div>
                <div class="label">File lỗi (không xử lý được)</div>
            </div>""", unsafe_allow_html=True)

        # ── Error details ────────────────────────────────────────────────────
        if errors:
            st.markdown("<br>", unsafe_allow_html=True)
            with st.expander(f"⚠️ {len(errors)} file gặp lỗi (click để xem chi tiết)"):
                for fname, err in errors:
                    st.error(f"**{fname}**: {err}")

        # ── Preview updated summary ──────────────────────────────────────────
        st.markdown("---")
        st.markdown("### 📋 Preview bảng tổng hợp sau khi cập nhật")

        try:
            preview_df = pd.read_excel(
                io.BytesIO(updated_bytes),
                header=None,
            )
            # Show first 50 rows, limited columns
            st.dataframe(
                preview_df.head(50),
                use_container_width=True,
                hide_index=False,
            )
        except Exception as e:
            st.warning(f"Không preview được bảng tổng hợp: {e}")

        # ── Download button ──────────────────────────────────────────────────
        st.markdown("<br>", unsafe_allow_html=True)
        st.download_button(
            label="⬇️  Tải file tổng hợp đã cập nhật (.xlsx)",
            data=updated_bytes,
            file_name="BIEU_TONG_HOP_DA_CAP_NHAT.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

        st.success("🎉 Xử lý hoàn tất! Bấm nút trên để tải file về máy.")


if __name__ == "__main__":
    main()
